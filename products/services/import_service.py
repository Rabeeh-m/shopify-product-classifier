import csv
import io
import logging
import os
import re

from django.conf import settings
from django.db import transaction
from django.utils import timezone as tz

from products.models import Product, ProductImage, ProductImport

logger = logging.getLogger(__name__)

REQUIRED_COLUMNS = {"title"}
OPTIONAL_COLUMNS = {"description", "brand", "product_type", "image_urls"}
ALL_COLUMNS = REQUIRED_COLUMNS | OPTIONAL_COLUMNS
IMAGE_SEPARATORS = [",", "|"]

# MIME types we accept per extension — used as a secondary check after extension.
_MIME_TYPES = {
    ".csv": {"text/csv", "application/vnd.ms-excel", "application/octet-stream"},
    ".xlsx": {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/zip",
        "application/octet-stream",
    },
}


class ParseError(Exception):
    """Raised when file parsing or validation fails during product import.

    Stores the list of individual error messages so the API can return
    them to the caller without losing detail.
    """

    def __init__(self, errors):
        self.errors = errors
        super().__init__(f"Parse failed: {'; '.join(errors)}")


def _sanitize_filename(filename):
    """Strip path components and dangerous characters from the filename.

    Only allows alphanumeric, hyphens, underscores, and dots.
    """
    basename = os.path.basename(filename or "upload")
    basename = re.sub(r"[^\w.\-]", "_", basename)
    basename = re.sub(r"_+", "_", basename).strip("_.")
    return basename or "upload"


def _normalize_header(col):
    """Lowercase, strip whitespace, and replace spaces with underscores.

    Handles None input by returning an empty string.
    """
    if col is None:
        return ""
    return col.strip().lower().replace(" ", "_")


def _validate_file(file_obj, filename):
    """Validate file extension, MIME type, and size.

    Checks:
    1. Extension is in ALLOWED_UPLOAD_EXTENSIONS.
    2. Declared content_type matches the expected MIME set for that extension.
    3. File size is within MAX_UPLOAD_SIZE_MB.
    """
    ext = os.path.splitext(filename)[1].lower()
    if ext not in settings.ALLOWED_UPLOAD_EXTENSIONS:
        raise ParseError(
            [
                f"Unsupported file type '{ext}'. "
                f"Allowed: {', '.join(sorted(settings.ALLOWED_UPLOAD_EXTENSIONS))}"
            ]
        )

    content_type = getattr(file_obj, "content_type", None)
    if content_type and ext in _MIME_TYPES:
        if content_type not in _MIME_TYPES[ext]:
            raise ParseError(
                [
                    f"Unexpected content type '{content_type}' for a {ext} file. "
                    f"Expected one of: {', '.join(sorted(_MIME_TYPES[ext]))}"
                ]
            )

    max_bytes = settings.MAX_UPLOAD_SIZE_MB * 1024 * 1024
    file_obj.seek(0, os.SEEK_END)
    size = file_obj.tell()
    file_obj.seek(0)
    if size > max_bytes:
        raise ParseError(
            [f"File size {size} bytes exceeds " f"maximum of {max_bytes} bytes"]
        )


def _read_csv(file_obj):
    """Read a CSV file and return (headers, rows) with normalized headers.

    Uses utf-8-sig encoding to handle BOM-prefixed files from Excel exports.
    """
    text = file_obj.read().decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    headers = [_normalize_header(h) for h in reader.fieldnames or []]
    rows = []
    for row in reader:
        normalized = {_normalize_header(k): v for k, v in row.items()}
        rows.append(normalized)
    return headers, rows


def _read_xlsx(file_obj):
    """Read an Excel .xlsx file and return (headers, rows) with normalized headers.

    Uses openpyxl in read-only mode for memory efficiency with large files.
    """
    from openpyxl import load_workbook

    try:
        wb = load_workbook(file_obj, read_only=True, data_only=True)
    except Exception as exc:
        raise ParseError([f"File is not a valid Excel (.xlsx) file: {exc}"]) from exc
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    raw_headers = next(rows_iter, [])
    headers = [_normalize_header(str(h)) if h else "" for h in raw_headers]
    rows = []
    for row in rows_iter:
        normalized = {}
        for i, val in enumerate(row):
            if i < len(headers):
                normalized[headers[i]] = str(val) if val is not None else ""
        rows.append(normalized)
    wb.close()
    return headers, rows


def _parse_image_urls(raw_value):
    """Parse a raw image URLs string into a list of individual URLs.

    Supports comma-separated and pipe-separated URLs. Returns an empty
    list for empty or whitespace-only input.
    """
    if not raw_value or not str(raw_value).strip():
        return []
    text = str(raw_value).strip()
    for sep in IMAGE_SEPARATORS:
        if sep in text:
            return [u.strip() for u in text.split(sep) if u.strip()]
    return [text]


def _validate_headers(headers):
    """Check that required columns are present and log unknown columns.

    Returns a list of error strings (empty if all required columns exist).
    """
    errors = []
    normalized = set(headers)
    missing = REQUIRED_COLUMNS - normalized
    if missing:
        errors.append(f"Missing required column(s): {', '.join(sorted(missing))}")
    unknown = normalized - ALL_COLUMNS
    if unknown:
        logger.warning("Skipping unknown columns: %s", sorted(unknown))
    return errors


def _create_products(rows, import_obj):
    """Create Product and ProductImage objects from parsed rows.

    Each product is created in its own transaction so a failure on one
    row doesn't roll back previously created products. Returns a tuple
    of (imported_count, failed_count, error_list).
    """
    imported = 0
    failed = 0
    errors = []

    for i, row in enumerate(rows, start=2):
        title = (row.get("title") or "").strip()
        if not title:
            failed += 1
            errors.append({"row": i, "error": "Missing required field: title"})
            continue

        with transaction.atomic():
            product = Product.objects.create(
                title=title,
                description=(row.get("description") or "").strip(),
                brand=(row.get("brand") or "").strip(),
                product_type=(row.get("product_type") or "").strip(),
                raw_data=row,
            )
            image_urls = _parse_image_urls(row.get("image_urls"))
            for url in image_urls:
                ProductImage.objects.create(product=product, url=url)
            imported += 1

    return imported, failed, errors


def import_products(file_obj, filename):
    """Validate, parse, and import products from a CSV or XLSX file.

    Creates a ProductImport record to track progress. Products with a
    missing title are counted as failures but don't halt the import.
    Raises ParseError on validation or header errors before any products
    are created.
    """
    _validate_file(file_obj, filename)
    file_obj.seek(0)
    ext = os.path.splitext(filename)[1].lower()
    if ext == ".csv":
        headers, rows = _read_csv(file_obj)
    else:
        headers, rows = _read_xlsx(file_obj)

    header_errors = _validate_headers(headers)
    if header_errors:
        raise ParseError(header_errors)

    import_obj = ProductImport.objects.create(
        file=file_obj,
        status=ProductImport.Status.PROCESSING,
        total_rows=len(rows),
    )

    try:
        imported, failed, row_errors = _create_products(rows, import_obj)
        import_obj.status = ProductImport.Status.COMPLETED
        import_obj.imported_rows = imported
        import_obj.failed_rows = failed
        import_obj.error_log = row_errors
        import_obj.completed_at = tz.now()
        import_obj.save(
            update_fields=[
                "status",
                "imported_rows",
                "failed_rows",
                "error_log",
                "completed_at",
            ]
        )
    except Exception as exc:
        logger.exception("Import failed for %s", filename)
        import_obj.status = ProductImport.Status.FAILED
        import_obj.error_log = [{"error": str(exc)}]
        import_obj.save(update_fields=["status", "error_log"])
        raise

    return import_obj
